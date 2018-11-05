import json
import re
import openpyxl
import traceback

"""
输入文件：data/0827话费流量类数据1.xlsx
输入文件格式：
A                        B                  C                   D                   E
端口号(无关内容)      模板主题(无关内容)      短信原文           模板内容(无关内容)        省份(无关内容)

配置文件：'data/kre.json'，'data/kre_list2.json' 在同级目录中

输出文件:
levelResult.txt

"""




class labeling():
    def __init__(self, ln):
        self.targetKey = ln
        self.isMatch = False
        self.result = []

    def match_newGeneralWithLoop(self, matched):

        idx_list = []

        for x in self.targetKey:
            idx_list.append(matched.start(x))
            idx_list.append(matched.end(x))

        bias = matched.start("stream")
        whole = matched.group("stream")

        re_str = ""
        count = 0
        sign = 0

        start = idx_list[count]
        end = idx_list[count + 1]

        value = ""
        for i, c in enumerate(whole):
            if i + bias < start or i + bias >= end:
                re_str += c
            else:
                if sign == 0:
                    re_str += c + ' B-X'
                    value += c
                    sign = 1
                    self.isMatch = True
                else:
                    re_str += c + ' I-X'
                    value += c
                continue
            if i + bias == end:
                self.result.append(value)
                value = ""
                sign = 0
                count += 2
                if count < 2 * len(self.targetKey):
                    start = idx_list[count]
                    end = idx_list[count + 1]
        if value != "":
            self.result.append(value)

        return re_str


def testReAndRecord(temp, line, l):
    af_line = re.sub(
        temp,
        l.match_newGeneralWithLoop,
        line.strip())

    if l.isMatch == True:
        if l.result == []:
            print("不对")
        return l.result
    else:
        return None


def calculateSegment(line):
    num = 0
    result = re.split("，|。|！|：|；", line)
    for x in result:
        if re.search("\d+", x) != None:
            num += 1
    return num


with open('data/kre.json', 'r') as fp:
    o = open("levelResult.txt", "w")

    fileJson = json.load(fp)

    records = fileJson['RECORDS']

    wb = openpyxl.load_workbook("data/0827话费流量类数据1.xlsx")
    sheet = wb["流量类"]

    max_row = sheet.max_row

    with open('data/kre_list2.json', 'r') as sp_fp:
        sp_fileJson = json.load(sp_fp)
        sp_records = sp_fileJson['RECORDS']
    print("记录一有%d个正则表达式" % len(records))
    print("记录二有%d个正则表达式" % len(sp_records))
    for row in range(2, max_row):

        line = str(sheet.cell(row=row, column=3).value).strip()

        num = 0

        key_value = {}

        for regex in records:
            temp_key = regex["key"]
            if regex["key"][:6] == "jsheng":
                my_list = [i.start() for i in re.finditer("_", regex["key"])]
                if len(my_list) == 2:
                    temp_key = temp_key[my_list[0] + 1:my_list[1]]

            l = labeling([temp_key])
            temp = "(?P<stream>" + regex["re"] + ")"

            result = testReAndRecord(temp, line, l)
            if result != None:
                if regex["cn_key"] not in key_value:
                    num += 1
                key_value[regex["cn_key"]] = result

        for regex in sp_records:
            temp_key_list = regex["cn_key"].split(" ")
            key_list = []
            for i, x in enumerate(temp_key_list):
                key_list.append("item" + str(i + 1))
            l = labeling(key_list)

            temp = "(?P<stream>" + regex["re"] + ")"

            try:
                result = testReAndRecord(temp, line, l)
            except:
                print(line)
                traceback.print_exc()
            if result != None:
                for i, x in enumerate(temp_key_list):
                    if x not in key_value:
                        num += 1
                    key_value[x] = [result[i]]

        total = calculateSegment(line)

        level = float(num) / (total + 0.00001)
        if level > 1:
            level = 1
        o.write(line)
        o.write("\t")
        o.write(str(level))
        o.write("\n")
        for x in key_value:
            o.write(x)
            o.write("\t")
            for y in key_value[x]:
                o.write(y)
                o.write("\t")
            o.write("\n")
        o.write("#####################")
        o.write("\n")
        o.flush()

    o.close()
