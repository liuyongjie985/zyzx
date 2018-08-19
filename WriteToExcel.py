# coding:utf-8

import xlsxwriter
from openpyxl import load_workbook

origin_out = []

my_list = ["./data/test.xlsx"]

for x in my_list:

    workbook = load_workbook(x)
    # booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
    sheets = workbook.get_sheet_names()  # 从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])

    rows = booksheet.rows
    columns = booksheet.columns
    # 迭代所有的行
    count = 0
    for row in rows:
        count += 1
        line = [col.value for col in row]
        temp = []
        if len(line) == 2:
            origin_out.append([line[0], line[1]])
    workbook.close()

workbook = xlsxwriter.Workbook('output/result.xlsx')

worksheet = workbook.add_worksheet()
worksheet.set_column("A:A", 20)


def writeExcel(row=0, port='端口', origin_sentence='原句', template='模板', kv='key_value对'):
    if row == 0:
        worksheet.write(row, 0, port)
        worksheet.write(row, 1, origin_sentence)
        worksheet.write(row, 2, template)
        worksheet.write(row, 3, kv)
    else:
        worksheet.write(row, 0, origin_out[idx][0])
        worksheet.write(row, 1, origin_out[idx][1])
        worksheet.write(row, 2, out[1][out[1].find(u'|') + 1:])
        c = 3
        for x in out[2][1:-2].strip().split(ur' '):
            if x == "":
                continue
            worksheet.write(row, c, x)
            c += 1


f = open("output/result_out")

sign = 0
out = []
row = 0
idx = 0

while 1:
    lines = f.readlines(100000)
    if not lines:
        break
    for line in lines:
        if line.strip() == "":
            continue

        out.append(line.decode('utf-8'))
        if len(out) == 3:
            writeExcel(row=row, port=origin_out[idx][0], origin_sentence=origin_out[idx][1],
                       template=out[1], kv=out[2])
            out = []
            row += 1
            idx += 1
workbook.close()
print row
