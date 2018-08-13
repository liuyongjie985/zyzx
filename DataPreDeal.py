# coding:utf-8

import re
from openpyxl import load_workbook
import xlrd
import csv
import traceback

# result = []


# 输出
o = open("data/TotalData.txt", "w")

my_list = ["./data/001-浙江.txt", "./data/002-内蒙古.txt", "./data/003-江苏短信10086.txt"]

for x in my_list:
    f = open(x)

    while 1:
        lines = f.readlines(100000)
        if not lines:
            break
        for line in lines:

            list = [i.start() for i in re.finditer(r'\|', line)]
            if len(list) == 2:
                o.write(line[:list[1]] + '|' + line[list[1] + 1:].strip() + "#")
                o.write('\n')
            # o.write(line.strip() + "#")
            # o.write()
    f.close()

my_list = ["./data/10086短信原文.xlsx", "./data/hangye30.xlsx"]

for x in my_list:

    workbook = load_workbook(x)
    # booksheet = workbook.active                #获取当前活跃的sheet,默认是第一个sheet
    sheets = workbook.get_sheet_names()  # 从名称获取sheet
    booksheet = workbook.get_sheet_by_name(sheets[0])

    rows = booksheet.rows
    columns = booksheet.columns
    # 迭代所有的行
    for row in rows:
        line = [col.value for col in row]
        print line
        try:
            if len(line) == 3:
                print line[2]
                o.write(str(line[1]).strip() + '|' + str(line[2]).strip() + "#")
                o.write('\n')

        except:
            traceback.print_exc()
            print line[2]
            try:
                o.write(str(line[1]).strip() + '|' + str(line[2].encode('utf-8')).strip() + "#")
            except:
                o.write(str(line[1].encode('utf-8')).strip() + '|' + str(line[2].encode('utf-8')).strip() + "#")
            o.write('\n')
        try:
            if len(line) == 2:
                print line[1]
                o.write(str(line[0]).strip() + '|' + str(line[1]).strip() + "#")
                o.write('\n')
        except:
            traceback.print_exc()
            print line[1]
            o.write(str(line[0].encode('utf-8')).strip() + '|' + str(line[1].encode('utf-8')).strip() + "#")
            o.write('\n')
    workbook.close()
# 通过坐标读取值

cell_11 = booksheet.cell(row=1, column=1).value

my_list = ["./data/hangye30_utf-8.txt"]

for x in my_list:
    f = open(x)
    while 1:
        lines = f.readlines(100000)
        if not lines:
            break
        for line in lines:
            o.write(line.strip() + "#")
            o.write('\n')

    f.close()

# 读取csv

csv_reader = csv.reader(open("data/hn.csv"))
for row in csv_reader:
    if len(row) == 2:
        o.write(row[0].strip() + '|' + row[1].strip() + "#")
        o.write('\n')
        # print row[1]

# 读取gbk文件
f = open("data/SMSLOG20160424.txt")
error_count = 0
total = 0
while 1:
    lines = f.readlines(100000)
    if not lines:
        break
    for line in lines:
        try:
            temp = line.decode('gbk')
            o.write(temp.encode('utf-8').strip() + "#")
            o.write('\n')
        except:
            error_count += 1
        total += 1
print "错误数:" + str(error_count)
print "总数：" + str(total)
print "未转化率：" + str(float(error_count) / total)
f.close()

f = open("data/sms_sample.txt")

while 1:
    lines = f.readlines(100000)
    if not lines:
        break
    for line in lines:
        o.write(line.strip() + "#")
        o.write('\n')
        # print line
f.close()

o.close()
